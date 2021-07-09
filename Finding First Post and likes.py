import json
import openpyxl
from openpyxl import Workbook
import pandas as pd

excel_file = openpyxl.load_workbook('Nifty50.xlsx')
companies = []
sheet = excel_file['Sheet1']

for x in range(3, 43):
    companies.append(sheet.cell(row=x, column=4).value)

excel_file.close()

wb = Workbook()
xlsheet = wb.active
xlsheet['A1'] = "Companies"
xlsheet['B1'] = "Avg Likes"
xlsheet['C1'] = "Avg Comments"
xlsheet['D1'] = "First Post"

i = 0
while i < len(companies):

    file = open('{}.json'.format(companies[i]), 'r')
    df = pd.read_json(file)
    # df.to_excel('airtelindia.xlsx')

    avg_likes = df['likes'].mean().round(2)
    avg_cmnts = df['comments'].mean().round(2)

    xlsheet.cell(row=i + 2, column=1).value = companies[i]
    xlsheet.cell(row=i + 2, column=2).value = avg_likes
    xlsheet.cell(row=i + 2, column=3).value = avg_cmnts

    date = df['DateOfPost'].iloc[-1]
    xlsheet.cell(row=i + 2, column=4).value = date[0:10]

    i += 1

wb.save('Comparison.xlsx')